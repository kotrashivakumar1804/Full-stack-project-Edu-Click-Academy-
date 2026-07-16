import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class NumberedCanvas(canvas.Canvas):
    """Custom canvas to compute total page count dynamically and print footers."""
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super(NumberedCanvas, self).showPage()
        super(NumberedCanvas, self).save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#64748b"))
        
        # Draw header
        self.drawString(54, 750 if self._pagesize == letter else 550, "Student Course Management Portal - Admin Reports")
        self.setStrokeColor(colors.HexColor("#e2e8f0"))
        self.setLineWidth(0.5)
        self.line(54, 742 if self._pagesize == letter else 542, self._pagesize[0] - 54, 742 if self._pagesize == letter else 542)
        
        # Draw footer
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(self._pagesize[0] - 54, 36, page_text)
        self.drawString(54, 36, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.restoreState()


def generate_pdf_report(title, headers, data, landscape_mode=False):
    """Generates a styled PDF report using ReportLab."""
    buffer = io.BytesIO()
    page_size = landscape(letter) if landscape_mode else letter
    
    # Bottom margin set to 54pt to leave room for footer
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=54, 
        rightMargin=54, 
        topMargin=72, 
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=12
    )
    
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#334155")
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10))
    
    # Prepare Table Data
    table_data = []
    # Add Header Row
    table_data.append([Paragraph(h, header_style) for h in headers])
    
    # Add Data Rows
    for row in data:
        row_cells = []
        for cell in row:
            val = str(cell) if cell is not None else ""
            row_cells.append(Paragraph(val, cell_style))
        table_data.append(row_cells)
        
    # Calculate column widths
    usable_width = page_size[0] - 108  # page size minus left and right margins
    num_cols = len(headers)
    col_width = usable_width / num_cols
    
    # Custom widths for specific tables if necessary
    col_widths = [col_width] * num_cols
    
    report_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Style Table
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ])
    
    # Add alternating row background colors
    for i in range(1, len(table_data)):
        bg = colors.HexColor("#f8fafc") if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        
    report_table.setStyle(t_style)
    story.append(report_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer


def generate_excel_report(sheet_name, headers, data):
    """Generates a styled Excel sheet using OpenPyXL."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Styling Palette
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    title_font = Font(name=font_family, size=16, bold=True, color="1E293B")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Write Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"{sheet_name} Report"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    # Spacer Row
    ws.append([])
    
    # Write Headers
    ws.append(headers)
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 25
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        
    # Write Data
    for row_data in data:
        # Format dates/numbers if needed before writing
        formatted_row = []
        for val in row_data:
            if isinstance(val, datetime):
                formatted_row.append(val.strftime("%Y-%m-%d %H:%M"))
            elif hasattr(val, 'strftime'):  # Date objects
                formatted_row.append(val.strftime("%Y-%m-%d"))
            else:
                formatted_row.append(val)
        ws.append(formatted_row)
        
    # Apply cell borders & zebra striping
    start_row = 4
    end_row = ws.max_row
    
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 20
        # Alternating background fill
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if r % 2 == 0 else PatternFill(fill_type=None)
        
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_fill.fill_type:
                cell.fill = row_fill
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't size column based on title row (row 1)
        for cell in col[1:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
